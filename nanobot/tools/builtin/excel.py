"""Spreadsheet reader and analytics tools (Excel + CSV)."""

from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import TYPE_CHECKING, Any, ClassVar

from nanobot.tools.base import Tool, ToolResult
from nanobot.tools.builtin.filesystem import _resolve_path

if TYPE_CHECKING:
    from nanobot.tools.result_cache import ToolResultCache

# Pattern for repeating date-allocation columns like "3/1/2026 (d)"
_DATE_ALLOC_RE = re.compile(r"^\d{1,2}/\d{1,2}/\d{4}\s*\(")

_EXCEL_SUFFIXES = {".xlsx", ".xls", ".xlsm", ".xlsb"}
_SUPPORTED_SUFFIXES = _EXCEL_SUFFIXES | {".csv"}

# Row threshold above which the tool adds a hint to use query_data / describe_data
_LARGE_DATASET_ROWS = 50

# Large text values are represented as cache-backed references instead of being
# inlined into tool output. This keeps the model from ingesting prompt-like data
# embedded in spreadsheet cells while still exposing the field size and a way to
# inspect the raw content.
_INLINE_VALUE_MAX_CHARS = 1000
_INLINE_PREVIEW_CHARS = 280


class ReadSpreadsheetTool(Tool):
    """Read an Excel workbook or CSV file and return structured data.

    Each sheet is cached individually so downstream tools (``excel_get_rows``,
    ``query_data``, ``describe_data``) can retrieve rows without the entire
    workbook needing to fit in a single cache entry.  The tool returns a
    compact metadata envelope (sheet names, headers, row counts, per-sheet
    cache keys) — never raw row data.
    """

    readonly = True
    # Per-sheet caching is handled internally; the registry should not
    # auto-cache the (small) metadata output again.
    cacheable = False

    def __init__(
        self,
        workspace: Path | None = None,
        allowed_dir: Path | None = None,
        max_rows: int = 200,
        cache: ToolResultCache | None = None,
    ):
        self._workspace = workspace
        self._allowed_dir = allowed_dir
        self._max_rows = max_rows
        self._cache: ToolResultCache | None = cache

    name = "read_spreadsheet"
    description = (
        "Read an Excel (.xlsx/.xls/.xlsm/.xlsb) workbook or CSV file and return "
        "sheet names, headers, and row data as JSON. Use this instead of exec for "
        "spreadsheet analysis. For large datasets, use describe_data or query_data "
        "with the returned cache_key for efficient analysis."
    )
    parameters: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "path": {
                "type": "string",
                "description": "Path to the file (.xlsx, .xls, .xlsm, .xlsb, or .csv)",
            },
            "sheet": {
                "type": "string",
                "description": ("Sheet name to read. If omitted, reads all sheets."),
            },
            "max_rows": {
                "type": "integer",
                "description": "Maximum rows to return per sheet (default 200)",
                "minimum": 1,
                "maximum": 5000,
            },
            "columns": {
                "type": "array",
                "description": ("Column names to include. If omitted, all columns are returned."),
                "items": {"type": "string"},
            },
        },
        "required": ["path"],
    }

    async def execute(  # type: ignore[override]
        self,
        path: str,
        sheet: str | None = None,
        max_rows: int | None = None,
        columns: list[str] | None = None,
        **kwargs: Any,
    ) -> ToolResult:
        try:
            file_path = _resolve_path(str(path), self._workspace, self._allowed_dir)
        except PermissionError as exc:
            return ToolResult.fail(f"Error: {exc}")

        if not file_path.exists():
            return ToolResult.fail(f"Error: File not found: {path}")
        if not file_path.is_file():
            return ToolResult.fail(f"Error: Not a file: {path}")

        suffix = file_path.suffix.lower()
        if suffix not in _SUPPORTED_SUFFIXES:
            return ToolResult.fail(f"Error: Unsupported file type: {suffix}")

        cap = max_rows or self._max_rows

        if suffix == ".csv":
            return self._read_csv(file_path, cap, columns)

        return self._read_excel(file_path, sheet, cap, columns)

    # ------------------------------------------------------------------
    # CSV reading
    # ------------------------------------------------------------------

    def _read_csv(
        self,
        file_path: Path,
        cap: int,
        columns: list[str] | None,
    ) -> ToolResult:
        try:
            raw_bytes = file_path.read_bytes()
        except Exception as exc:
            return ToolResult.fail(f"Error reading file: {exc}")

        text = raw_bytes.decode("utf-8-sig", errors="replace")

        # Detect delimiter via csv.Sniffer (fall back to comma)
        try:
            sample = text[:8192]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel  # type: ignore[assignment]

        reader = csv.reader(io.StringIO(text), dialect)
        rows_iter = iter(reader)

        # Skip leading blank rows (e.g. files starting with an empty line)
        header_row = next(rows_iter, None)
        while header_row is not None and not header_row:
            header_row = next(rows_iter, None)

        if header_row is None:
            sname = file_path.stem
            empty_data: dict[str, Any] = {"headers": [], "rows": [], "total_rows": 0}
            cache_key = self._cache_sheet(str(file_path), sname, empty_data)
            empty_meta = {
                "headers": [],
                "row_count": 0,
                "total_rows": 0,
                "cache_key": cache_key,
            }
            return ToolResult.ok(
                json.dumps(
                    {"file": file_path.name, "sheets": {sname: empty_meta}},
                    ensure_ascii=False,
                )
            )

        # CSV rows are always strings — convert to native types below
        raw_rows: list[list[Any]] = []
        total = 0
        for row in rows_iter:
            total += 1
            if total <= cap:
                raw_rows.append([_coerce_csv_value(v) for v in row])

        sheet_data = _build_sheet_data(
            header_row,
            raw_rows,
            total,
            cap,
            columns,
        )

        sname = file_path.stem
        cache_key = self._cache_sheet(str(file_path), sname, sheet_data)
        meta = _sheet_meta(sheet_data, cache_key)
        result: dict[str, Any] = {"file": file_path.name, "sheets": {sname: meta}}
        return ToolResult.ok(json.dumps(result, ensure_ascii=False, default=str))

    # ------------------------------------------------------------------
    # Excel reading
    # ------------------------------------------------------------------

    def _read_excel(
        self,
        file_path: Path,
        sheet: str | None,
        cap: int,
        columns: list[str] | None,
    ) -> ToolResult:
        try:
            import openpyxl
        except ImportError:
            return ToolResult.fail(
                "openpyxl is not available. Excel files (.xlsx) cannot be read. "
                "Only CSV files are supported in this environment.",
                error_type="missing_dependency",
            )

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:
            return ToolResult.fail(f"Error opening workbook: {exc}")

        try:
            sheet_names = wb.sheetnames
            target_sheets = [sheet] if sheet else sheet_names

            if sheet and sheet not in sheet_names:
                wb.close()
                return ToolResult.fail(
                    f"Sheet '{sheet}' not found. Available: {', '.join(sheet_names)}"
                )

            result: dict[str, Any] = {"file": str(file_path.name), "sheets": {}}

            for sname in target_sheets:
                ws = wb[sname]
                rows_iter = ws.iter_rows(values_only=True)

                header_row = next(rows_iter, None)
                if header_row is None:
                    cache_key = self._cache_sheet(
                        str(file_path),
                        sname,
                        {"headers": [], "rows": [], "total_rows": 0},
                    )
                    result["sheets"][sname] = {
                        "headers": [],
                        "row_count": 0,
                        "total_rows": 0,
                        "cache_key": cache_key,
                    }
                    continue

                headers = list(header_row)

                raw_rows: list[list[Any]] = []
                total = 0
                for row in rows_iter:
                    total += 1
                    if total <= cap:
                        raw_rows.append(list(row))

                sheet_data = _build_sheet_data(
                    headers,
                    raw_rows,
                    total,
                    cap,
                    columns,
                )
                cache_key = self._cache_sheet(str(file_path), sname, sheet_data)
                result["sheets"][sname] = _sheet_meta(sheet_data, cache_key)

            wb.close()
        except Exception as exc:
            wb.close()
            return ToolResult.fail(f"Error reading workbook: {exc}")

        return ToolResult.ok(json.dumps(result, ensure_ascii=False, default=str))

    # ------------------------------------------------------------------
    # Per-sheet cache helpers
    # ------------------------------------------------------------------

    def _cache_sheet(
        self,
        file_path: str,
        sheet_name: str,
        sheet_data: dict[str, Any],
    ) -> str:
        """Store a single sheet's data in the cache and return the cache key."""
        if self._cache is None:
            return ""
        full_json = json.dumps(
            {
                "headers": sheet_data.get("headers", []),
                "rows": sheet_data.get("rows", []),
                "row_count": sheet_data.get("row_count", len(sheet_data.get("rows", []))),
                "total_rows": sheet_data.get("total_rows", 0),
            },
            ensure_ascii=False,
            default=str,
        )
        return self._cache.store(
            "read_spreadsheet",
            {"path": file_path, "sheet": sheet_name},
            full_json,
            "",  # summary left empty — LLM sees the metadata envelope instead
            token_estimate=len(full_json) // 4,
        )


def _sheet_meta(sheet_data: dict[str, Any], cache_key: str) -> dict[str, Any]:
    """Build a compact metadata dict for one sheet (no raw rows)."""
    meta: dict[str, Any] = {
        "headers": sheet_data.get("headers", []),
        "row_count": sheet_data.get("row_count", 0),
        "total_rows": sheet_data.get("total_rows", 0),
        "cache_key": cache_key,
    }
    if sheet_data.get("truncated"):
        meta["truncated"] = True
        if sheet_data.get("note"):
            meta["note"] = sheet_data["note"]
    if sheet_data.get("empty_columns_removed"):
        meta["empty_columns_removed"] = sheet_data["empty_columns_removed"]
    if sheet_data.get("date_columns_collapsed"):
        meta["date_columns_collapsed"] = sheet_data["date_columns_collapsed"]
    if sheet_data.get("hint"):
        meta["hint"] = sheet_data["hint"]
    return meta


def _coerce_csv_value(val: str) -> Any:
    """Attempt to coerce a CSV string to a native Python type."""
    if not val:
        return None
    # Integer
    try:
        return int(val)
    except ValueError:
        pass
    # Float
    try:
        return float(val)
    except ValueError:
        pass
    return val


def _build_sheet_data(
    raw_headers: list[Any],
    raw_rows: list[list[Any]],
    total: int,
    cap: int,
    columns: list[str] | None,
) -> dict[str, Any]:
    """Build the per-sheet JSON dict from headers and raw row data.

    Shared by both the Excel and CSV code paths.  Handles column filtering,
    empty-column stripping, date-allocation collapsing, and empty-row skip.
    """
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(raw_headers)]

    # Determine column indices to keep
    if columns:
        col_set = set(columns)
        keep_idx = [i for i, h in enumerate(headers) if h in col_set]
        if not keep_idx:
            return {
                "headers": headers,
                "rows": [],
                "total_rows": 0,
                "note": f"None of {columns} matched headers",
            }
        headers = [headers[i] for i in keep_idx]
        raw_rows = [[vals[i] if i < len(vals) else None for i in keep_idx] for vals in raw_rows]
    else:
        # Pad short rows to header length
        raw_rows = [
            vals + [None] * (len(headers) - len(vals)) if len(vals) < len(headers) else vals
            for vals in raw_rows
        ]

    # Auto-strip entirely empty columns
    stripped_count = 0
    if not columns and raw_rows:
        non_empty: set[int] = set()
        for vals in raw_rows:
            for ci, v in enumerate(vals):
                if v is not None and not (isinstance(v, str) and not v.strip()):
                    non_empty.add(ci)
        if len(non_empty) < len(headers):
            stripped_count = len(headers) - len(non_empty)
            ordered = sorted(non_empty)
            headers = [headers[ci] for ci in ordered]
            raw_rows = [[vals[ci] for ci in ordered] for vals in raw_rows]

    # Separate date-allocation columns from metadata columns
    meta_idx: list[int] = []
    alloc_idx: list[int] = []
    for ci, h in enumerate(headers):
        if _DATE_ALLOC_RE.match(h):
            alloc_idx.append(ci)
        else:
            meta_idx.append(ci)

    # Build row dicts, skipping empty rows
    data_rows: list[dict[str, Any]] = []
    for vals in raw_rows:
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in vals):
            continue
        record: dict[str, Any] = {}
        for ci in meta_idx:
            sv = _serialize(vals[ci])
            if sv is not None:
                record[headers[ci]] = sv
        if alloc_idx:
            alloc_summary = _summarize_allocations([(headers[ci], vals[ci]) for ci in alloc_idx])
            if alloc_summary:
                record["_allocations"] = alloc_summary
        if record:
            data_rows.append(record)

    meta_headers = [headers[ci] for ci in meta_idx]
    sheet_data: dict[str, Any] = {
        "headers": meta_headers,
        "row_count": len(data_rows),
        "total_rows": total,
        "rows": data_rows,
    }
    if total > cap:
        sheet_data["truncated"] = True
        sheet_data["note"] = f"Showing first {cap} of {total} rows"
    if stripped_count:
        sheet_data["empty_columns_removed"] = stripped_count
    if alloc_idx:
        sheet_data["date_columns_collapsed"] = len(alloc_idx)
    if len(data_rows) > _LARGE_DATASET_ROWS:
        sheet_data["hint"] = (
            "For analysis, use describe_data or query_data with this data's cache_key"
        )
    return sheet_data


def _serialize(val: Any) -> Any:
    """Convert cell value to a JSON-friendly type."""
    if val is None:
        return None
    # datetime objects → ISO string
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return val


def _cache_large_text(
    cache: ToolResultCache | None,
    *,
    source_cache_key: str,
    row_index: int,
    column: str,
    text: str,
) -> str | None:
    """Store a large cell value in the tool cache and return its cache key."""
    if cache is None:
        return None
    return cache.store(
        "spreadsheet_large_text",
        {
            "source_cache_key": source_cache_key,
            "row_index": row_index,
            "column": column,
        },
        text,
        "",
        token_estimate=len(text) // 4,
    )


def _sanitize_cell_value(
    value: Any,
    *,
    cache: ToolResultCache | None,
    source_cache_key: str,
    row_index: int,
    column: str,
) -> Any:
    """Replace oversized string values with explicit cache-backed references."""
    if not isinstance(value, str) or len(value) <= _INLINE_VALUE_MAX_CHARS:
        return value

    cached_key = _cache_large_text(
        cache,
        source_cache_key=source_cache_key,
        row_index=row_index,
        column=column,
        text=value,
    )
    preview = value[:_INLINE_PREVIEW_CHARS]
    ref: dict[str, Any] = {
        "kind": "large_text_ref",
        "length": len(value),
        "preview": preview,
        "truncated_preview": len(value) > len(preview),
        "note": "Untrusted cell data omitted from inline output; inspect via cache_get_slice.",
    }
    if cached_key:
        ref["cache_key"] = cached_key
    return ref


def _sanitize_row_for_output(
    row: dict[str, Any],
    *,
    cache: ToolResultCache | None,
    source_cache_key: str,
    row_index: int,
) -> tuple[dict[str, Any], bool]:
    """Sanitize a row for tool output, returning the row and whether it changed."""
    changed = False
    sanitized: dict[str, Any] = {}
    for column, value in row.items():
        sanitized_value = _sanitize_cell_value(
            value,
            cache=cache,
            source_cache_key=source_cache_key,
            row_index=row_index,
            column=column,
        )
        if sanitized_value is not value:
            changed = True
        sanitized[column] = sanitized_value
    return sanitized, changed


def _summarize_allocations(pairs: list[tuple[str, Any]]) -> str | None:
    """Collapse date-allocation column values into a compact summary.

    Given pairs of (header, value) for date columns like "3/1/2026 (d)",
    returns a string like "3/3/2026..12/4/2026 total=350.0 avg=2.0/day"
    or None if no non-null values exist.
    """
    non_null = [(h, v) for h, v in pairs if v is not None and v != 0]
    if not non_null:
        return None
    first_date = non_null[0][0].split("(")[0].strip()
    last_date = non_null[-1][0].split("(")[0].strip()
    total = sum(float(v) for _, v in non_null if isinstance(v, int | float))
    days = len(non_null)
    avg = round(total / days, 2) if days else 0
    return f"{first_date}..{last_date} days={days} total={total} avg={avg}/day"


# ---------------------------------------------------------------------------
# SQL safety validation
# ---------------------------------------------------------------------------

# Statements that are allowed as the leading keyword in a query.
_ALLOWED_SQL_STARTS = {"select", "with"}

# Keywords that indicate a mutating / dangerous statement.
_DISALLOWED_SQL = re.compile(
    r"\b(insert|update|delete|drop|create|alter|truncate|exec|call|grant|revoke|"
    r"attach|detach|copy|load|install|export)\b",
    re.IGNORECASE,
)


def _validate_select_only(sql: str) -> str | None:
    """Return ``None`` if *sql* is a safe SELECT/WITH query, else an error message."""
    stripped = sql.strip()

    # Strip leading SQL comments
    while stripped.startswith("--"):
        stripped = stripped.split("\n", 1)[-1].strip()
    while stripped.startswith("/*"):
        end = stripped.find("*/")
        if end == -1:
            return "Unterminated block comment in SQL"
        stripped = stripped[end + 2 :].strip()

    if not stripped:
        return "Empty SQL query"

    first_word = re.split(r"\s", stripped, maxsplit=1)[0].lower().rstrip("(")
    if first_word not in _ALLOWED_SQL_STARTS:
        return f"Only SELECT queries are allowed (got '{first_word.upper()}')"

    # Reject stacked statements (e.g. "SELECT 1; DROP TABLE data")
    if ";" in stripped:
        return "Multiple statements are not allowed"

    # Reject embedded mutating keywords (catches injection attempts)
    if _DISALLOWED_SQL.search(stripped):
        return "Query contains disallowed keywords"

    return None


# ---------------------------------------------------------------------------
# Shared helper: extract rows from cached result
# ---------------------------------------------------------------------------


def _get_cached_sheet_rows(
    cache: ToolResultCache,
    cache_key: str,
    sheet: str | None,
) -> tuple[list[dict[str, Any]], str, str | None]:
    """Return ``(rows, sheet_name, error)`` from a cached spreadsheet result.

    Supports both per-sheet cache entries (Option C — rows live at the top
    level) and legacy whole-workbook entries (rows nested under ``sheets``).
    On success *error* is ``None``; on failure *rows* is empty and *error*
    contains the message to return to the caller.
    """
    entry = cache.get(cache_key)
    if entry is None:
        return [], "", f"No cached result for key '{cache_key}'."

    try:
        parsed = json.loads(entry.full_output)
    except (json.JSONDecodeError, TypeError):
        return [], "", "Cached result is not valid JSON."

    # Per-sheet format: rows at top level (no "sheets" wrapper)
    if "rows" in parsed and "sheets" not in parsed:
        rows: list[dict[str, Any]] = parsed.get("rows", [])
        return rows, sheet or "sheet", None

    # Legacy whole-workbook format
    sheets: dict[str, Any] = parsed.get("sheets", {})
    if not sheets:
        return [], "", "No sheets found in cached result."

    if sheet:
        sheet_data = sheets.get(sheet)
        if not sheet_data:
            available = list(sheets.keys())
            return [], "", f"Sheet '{sheet}' not found. Available: {available}"
        sname = sheet
    else:
        sname = next(iter(sheets))
        sheet_data = sheets[sname]

    rows = sheet_data.get("rows", [])
    return rows, sname, None


# ---------------------------------------------------------------------------
# DuckDB helpers
# ---------------------------------------------------------------------------


def _load_rows_to_duckdb(
    rows: list[dict[str, Any]],
) -> Any:
    """Create an in-memory DuckDB connection with *rows* loaded as table ``data``.

    Returns the open ``duckdb.DuckDBPyConnection``.  Caller must close it.
    """
    import tempfile

    import duckdb

    conn = duckdb.connect(":memory:")
    # DuckDB replacement scans don't accept plain list[dict].
    # Write JSON to a temp file and use read_json_auto on the path.
    rows_json = json.dumps(rows, ensure_ascii=False, default=str)
    tmp = tempfile.NamedTemporaryFile(
        mode="w",
        suffix=".json",
        delete=False,
        encoding="utf-8",
    )
    try:
        tmp.write(rows_json)
        tmp.close()
        conn.execute(
            "CREATE TABLE data AS SELECT * FROM read_json_auto(?)",
            [tmp.name],
        )
    finally:
        Path(tmp.name).unlink(missing_ok=True)
    return conn


# ---------------------------------------------------------------------------
# DuckDB-powered analytical tools
# ---------------------------------------------------------------------------


class QueryDataTool(Tool):
    """Run a SQL SELECT query over previously loaded spreadsheet data."""

    readonly = True

    def __init__(self, cache: ToolResultCache) -> None:
        self._cache = cache

    name = "query_data"
    description = (
        "Run a SQL SELECT query over previously loaded spreadsheet data. "
        "The data is available as a table named 'data'. "
        "Use the cache_key from a prior read_spreadsheet call. "
        "Example: SELECT department, SUM(amount) FROM data GROUP BY department"
    )
    parameters: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "cache_key": {
                "type": "string",
                "description": "The cache key from a prior read_spreadsheet call.",
            },
            "query": {
                "type": "string",
                "description": "SQL SELECT query. The table is named 'data'.",
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name to query. If omitted, queries the first sheet.",
            },
        },
        "required": ["cache_key", "query"],
    }

    async def execute(  # type: ignore[override]
        self,
        cache_key: str,
        query: str,
        sheet: str | None = None,
        **kwargs: Any,
    ) -> ToolResult:
        # Validate SQL safety
        error = _validate_select_only(query)
        if error:
            return ToolResult.fail(error, error_type="validation")

        rows, sname, err = _get_cached_sheet_rows(self._cache, cache_key, sheet)
        if err:
            return ToolResult.fail(err, error_type="not_found")
        if not rows:
            return ToolResult.fail(f"Sheet '{sname}' has no data rows.", error_type="not_found")

        try:
            conn = _load_rows_to_duckdb(rows)
            result = conn.execute(query)
            columns = [desc[0] for desc in result.description]
            fetched = result.fetchall()
            conn.close()
        except Exception as exc:
            return ToolResult.fail(f"SQL error: {exc}")

        all_rows = [dict(zip(columns, row)) for row in fetched]
        total_result_rows = len(all_rows)

        # Serialize rows until the output budget is exhausted
        result_rows: list[dict[str, Any]] = []
        budget = _MAX_OUTPUT_CHARS
        sanitized_fields = 0
        for row_index, row in enumerate(all_rows):
            safe_row, changed = _sanitize_row_for_output(
                row,
                cache=self._cache,
                source_cache_key=cache_key,
                row_index=row_index,
            )
            if changed:
                sanitized_fields += 1
            encoded = json.dumps(safe_row, ensure_ascii=False, default=str)
            budget -= len(encoded)
            if budget < 0 and result_rows:
                break
            result_rows.append(safe_row)

        result_data: dict[str, Any] = {
            "columns": columns,
            "row_count": total_result_rows,
            "returned": len(result_rows),
            "rows": result_rows,
        }
        if sanitized_fields:
            result_data["sanitized_rows"] = sanitized_fields
            result_data["note"] = (
                "Oversized text fields were replaced with cache-backed references to avoid "
                "injecting raw prompt-like data into the model context."
            )
        if len(result_rows) < total_result_rows:
            size_limit_note = (
                f"Showing {len(result_rows)} of {total_result_rows} result rows (output size limit). "
                "Refine your SQL with LIMIT/OFFSET or WHERE clauses to page through more."
            )
            if "note" in result_data:
                result_data["note"] += " " + size_limit_note
            else:
                result_data["note"] = size_limit_note

        output = json.dumps(result_data, ensure_ascii=False, default=str)
        return ToolResult.ok(output)


class DescribeDataTool(Tool):
    """Get schema, statistics, and sample rows from loaded spreadsheet data."""

    readonly = True

    def __init__(self, cache: ToolResultCache) -> None:
        self._cache = cache

    name = "describe_data"
    description = (
        "Get schema, statistics, and sample rows from previously loaded "
        "spreadsheet data. Returns column names, types, row count, null counts, "
        "unique counts, min/max for numerics, and sample rows. "
        "Use this before writing SQL queries with query_data."
    )
    parameters: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "cache_key": {
                "type": "string",
                "description": "The cache key from a prior read_spreadsheet call.",
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name. If omitted, describes the first sheet.",
            },
        },
        "required": ["cache_key"],
    }

    async def execute(  # type: ignore[override]
        self,
        cache_key: str,
        sheet: str | None = None,
        **kwargs: Any,
    ) -> ToolResult:
        rows, sname, err = _get_cached_sheet_rows(self._cache, cache_key, sheet)
        if err:
            return ToolResult.fail(err, error_type="not_found")
        if not rows:
            return ToolResult.fail(f"Sheet '{sname}' has no data rows.", error_type="not_found")

        try:
            conn = _load_rows_to_duckdb(rows)

            # Column info from DESCRIBE
            desc_result = conn.execute("DESCRIBE data").fetchall()
            col_info: list[dict[str, Any]] = []
            for col_name, col_type, *_ in desc_result:
                stats = conn.execute(
                    f'SELECT COUNT("{col_name}") AS non_null, '
                    f'COUNT(DISTINCT "{col_name}") AS "unique", '
                    f'MIN("{col_name}") AS min_val, '
                    f'MAX("{col_name}") AS max_val '
                    f"FROM data"
                ).fetchone()
                entry: dict[str, Any] = {
                    "name": col_name,
                    "type": str(col_type),
                }
                if stats:
                    entry["non_null"] = stats[0]
                    entry["unique"] = stats[1]
                    entry["min"] = _serialize(stats[2])
                    entry["max"] = _serialize(stats[3])
                col_info.append(entry)

            total_rows = conn.execute("SELECT COUNT(*) FROM data").fetchone()
            row_count = total_rows[0] if total_rows else 0

            sample = conn.execute("SELECT * FROM data LIMIT 5").fetchall()
            col_names = [c[0] for c in conn.execute("DESCRIBE data").fetchall()]
            sample_rows = [dict(zip(col_names, row)) for row in sample]

            conn.close()
        except Exception as exc:
            return ToolResult.fail(f"Error describing data: {exc}")

        output = json.dumps(
            {
                "sheet": sname,
                "row_count": row_count,
                "columns": col_info,
                "sample_rows": sample_rows,
            },
            ensure_ascii=False,
            default=str,
        )
        return ToolResult.ok(output)


# ---------------------------------------------------------------------------
# Cache-backed retrieval tools
# ---------------------------------------------------------------------------


# Output budget (chars) per excel_get_rows / excel_find / query_data call.
# Not truncation — the agent can page through with start_row/end_row.
# ~30 KB ≈ ~8K tokens — adapts naturally: simple rows → many, huge rows → few.
_MAX_OUTPUT_CHARS = 30_000


class ExcelGetRowsTool(Tool):
    """Retrieve a row range from a previously cached Excel result."""

    readonly = True
    cacheable = False

    def __init__(self, cache: ToolResultCache) -> None:
        self._cache = cache

    name = "excel_get_rows"
    description = (
        "Retrieve a range of rows from a previously cached read_spreadsheet result. "
        "Use the cache_key from the read_spreadsheet summary. "
        "Output is size-limited per call — use start_row/end_row to page through large data."
    )
    parameters: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "cache_key": {
                "type": "string",
                "description": "The cache key from a prior read_spreadsheet call.",
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name. If omitted, returns rows from the first sheet.",
            },
            "start_row": {
                "type": "integer",
                "description": "Start row index (0-based). Default 0.",
                "minimum": 0,
            },
            "end_row": {
                "type": "integer",
                "description": "End row index (exclusive). Default 25.",
                "minimum": 1,
            },
        },
        "required": ["cache_key"],
    }

    async def execute(  # type: ignore[override]
        self,
        cache_key: str,
        sheet: str | None = None,
        start_row: int = 0,
        end_row: int = 25,
        **kwargs: Any,
    ) -> ToolResult:
        rows, _sname, err = _get_cached_sheet_rows(self._cache, cache_key, sheet)
        if err:
            return ToolResult.fail(err, error_type="not_found")

        total = len(rows)
        actual_end = min(end_row, total)

        # Serialize rows until the output budget is exhausted
        included: list[dict[str, Any]] = []
        budget = _MAX_OUTPUT_CHARS
        sanitized_rows = 0
        for row_index in range(start_row, actual_end):
            safe_row, changed = _sanitize_row_for_output(
                rows[row_index],
                cache=self._cache,
                source_cache_key=cache_key,
                row_index=row_index,
            )
            if changed:
                sanitized_rows += 1
            encoded = json.dumps(safe_row, ensure_ascii=False, default=str)
            budget -= len(encoded)
            if budget < 0 and included:
                break
            included.append(safe_row)

        rows_end = start_row + len(included)
        result: dict[str, Any] = {
            "rows": included,
            "range": f"{start_row}-{rows_end}",
            "total_rows": total,
        }
        if sanitized_rows:
            result["sanitized_rows"] = sanitized_rows
            result["note"] = (
                "Oversized text fields were replaced with cache-backed references to avoid "
                "injecting raw prompt-like data into the model context."
            )
        if rows_end < total:
            result["next_page"] = f"start_row={rows_end}, end_row={min(rows_end + 25, total)}"
            result["remaining_rows"] = total - rows_end

        output = json.dumps(result, ensure_ascii=False, default=str)
        return ToolResult.ok(output)


class ExcelFindTool(Tool):
    """Search cached Excel data for matching rows."""

    readonly = True
    cacheable = False

    def __init__(self, cache: ToolResultCache) -> None:
        self._cache = cache

    name = "excel_find"
    description = (
        "Search a cached read_spreadsheet result for rows matching a text query. "
        "Searches all columns or a specific column."
    )
    parameters: ClassVar[dict[str, Any]] = {
        "type": "object",
        "properties": {
            "cache_key": {
                "type": "string",
                "description": "The cache key from a prior read_spreadsheet call.",
            },
            "query": {
                "type": "string",
                "description": "Text to search for (case-insensitive substring match).",
            },
            "column": {
                "type": "string",
                "description": "Column name to limit search to. If omitted, searches all.",
            },
            "sheet": {
                "type": "string",
                "description": "Sheet name. If omitted, searches the first sheet.",
            },
        },
        "required": ["cache_key", "query"],
    }

    async def execute(  # type: ignore[override]
        self,
        cache_key: str,
        query: str,
        column: str | None = None,
        sheet: str | None = None,
        **kwargs: Any,
    ) -> ToolResult:
        rows, _sname, err = _get_cached_sheet_rows(self._cache, cache_key, sheet)
        if err:
            return ToolResult.fail(err, error_type="not_found")

        q = query.lower()
        matches: list[tuple[int, dict[str, Any]]] = []

        for i, row in enumerate(rows):
            if column:
                val = row.get(column, "")
                if q in str(val).lower():
                    matches.append((i, row))
            else:
                if any(q in str(v).lower() for v in row.values()):
                    matches.append((i, row))

        # Serialize matches until the output budget is exhausted
        page: list[dict[str, Any]] = []
        budget = _MAX_OUTPUT_CHARS
        sanitized_rows = 0
        for idx, row in matches:
            safe_row, changed = _sanitize_row_for_output(
                row,
                cache=self._cache,
                source_cache_key=cache_key,
                row_index=idx,
            )
            if changed:
                sanitized_rows += 1
            entry = {"row_index": idx, **safe_row}
            encoded = json.dumps(entry, ensure_ascii=False, default=str)
            budget -= len(encoded)
            if budget < 0 and page:
                break
            page.append(entry)

        result: dict[str, Any] = {
            "query": query,
            "total_matches": len(matches),
            "returned": len(page),
            "rows": page,
        }
        if sanitized_rows:
            result["sanitized_rows"] = sanitized_rows
            result["note"] = (
                "Oversized text fields were replaced with cache-backed references to avoid "
                "injecting raw prompt-like data into the model context."
            )
        if len(page) < len(matches):
            size_limit_note = (
                f"Showing {len(page)} of {len(matches)} matches (output size limit). "
                "Use excel_get_rows with specific row indices to retrieve more."
            )
            if "note" in result:
                result["note"] += " " + size_limit_note
            else:
                result["note"] = size_limit_note

        output = json.dumps(result, ensure_ascii=False, default=str)
        return ToolResult.ok(output)
